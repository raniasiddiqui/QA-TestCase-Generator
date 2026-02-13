import streamlit as st
import asyncio
import re
import io
import nest_asyncio
from groq import Groq
from playwright.async_api import async_playwright
from urllib.parse import urlparse
from collections import deque
import requests
import pandas as pd
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import json
import sys

if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

nest_asyncio.apply()

# --- 1. ORIGINAL SCRIPT CODE (UNMODIFIED) ---
# Load config
# with open("config.json", "r") as f:
#     CONFIG = json.load(f)

# GROQ_API_KEY = CONFIG["groq_api_key"]
# DEFAULT_GROQ_MODEL = CONFIG["groq_default_model"]
GROQ_API_KEY = st.secrets["groq_api_key"]
DEFAULT_GROQ_MODEL = st.secrets["groq_default_model"]

# Initialize the Groq client
try:
    groq_client = Groq(api_key=GROQ_API_KEY)
except Exception as e:
    st.error(f"Failed to initialize Groq client. Check your API key in config.json. Error: {e}")
    st.stop()
# Light blue theme ‚Äì pick one shade you like
light_blue = "#a3d8ff"      # very light
# light_blue = "#81d4fa"    # a bit more vivid
# light_blue = "#4fc3f7"    # stronger but still friendly

st.markdown(
    f"""
    <style>
        button[kind="primary"] {{
            background-color: {light_blue} !important;
            border-color: {light_blue} !important;
            color: black !important;           /* or #111 for better contrast */
        }}
        button[kind="primary"]:hover {{
            background-color: #90caf9 !important;   /* slightly darker on hover */
            border-color: #90caf9 !important;
        }}
    </style>
    """,
    unsafe_allow_html=True
)
class GroqOSSAgent:
    """Base agent class using Groq models"""
    def __init__(self, name: str, system_message: str, model_name: str = DEFAULT_GROQ_MODEL):
        self.name = name
        self.system_message = system_message
        self.model_name = model_name

    async def generate_response(self, message: str) -> str:
        try:
            def run_completion():
                completion = groq_client.chat.completions.create(
                    model=self.model_name,
                    max_tokens=16384,
                    messages=[
                        {"role": "system", "content": self.system_message},
                        {"role": "user", "content": message}
                    ]
                )
                return completion.choices[0].message.content

            loop = asyncio.get_event_loop()
            return await loop.run_in_executor(None, run_completion)
        except Exception as e:
            return f"Error generating Groq OSS response: {str(e)}"

async def refine_instruction(instruction: str) -> str:
    refiner = GroqOSSAgent(
        name="InstructionRefiner",
        system_message="""
        You are an expert in writing clear, precise, and unambiguous instructions for QA automation tasks.
        Your task is to refine the provided instruction and make it understandable by an LLM easily, to ensure it is:
        - Clear and concise, actionable language, avoiding ambiguity.
        - Unambiguous with no vague terms
        - Structured for easy interpretation by automation agents
        - Focused on specifying exact actions, selectors, and validations
        - Compliant with Playwright sync API requirements
        - Includes self-healing locator guidelines
        - Avoids placeholders or vague instructions
        - Follow Playwright sync API conventions
        - Include possible self-healing locator strategies. These include ID, name, class name, tag name, CSS selector, XPath, and role-based selectors, and text-based selectors. They should be prioritized based on reliability and stability.
        - Focus on:
          - Setup steps (navigate, prepare data)
          - Action steps (click, fill, submit)
          - Verification steps (assertions, checks)
          - Error handling considerations         
        - Requests per-step pass/fail logging and assertions
        Output only the refined instruction as plain text, no markdown or explanations. Dont output any testcases in this step.
        """,
        model_name=DEFAULT_GROQ_MODEL
    )
    return await refiner.generate_response(instruction)

class SiteInspectorAgent(GroqOSSAgent):
    def __init__(self):
        system_message = """
        You are a site inspector that analyzes crawled web pages to extract reliable Playwright locators and discover QA-relevant insights for comprehensive test case generation.
        You receive snippets from multiple crawled pages of the site and the user's instruction describing specific functionalities.
        Analyze the crawled page snippets and user instruction to:
        - Summarize the site structure, key pages, navigation flows, and discovered features (e.g., forms, buttons, interactive elements, user journeys).
        - Identify possible test scenarios based on the site's elements and the user's instruction, including core functionalities, alternative flows, edge cases, and error conditions.
        - Extract and recommend reliable Playwright locators (ID, name, class name, tag name, CSS selector, XPath, role-based, text-based) for key elements mentioned in the instruction or discovered during crawling.
        - Suggest self-healing locator strategies and waits for dynamic content, prioritizing reliability and stability.
        - Provide insights to generate a wider range of test cases, such as alternative paths, error-prone areas, and integration points.
        Output a string starting with 'Site Insights and Recommended Locators: ' followed by a structured summary:
        - Site Structure: Summarize key pages, navigation patterns, and features.
        - Discovered Test Scenarios: List potential test cases (e.g., functional, negative, edge cases) based on crawled data and instruction.
        - Recommended Locators: List reliable locators for key elements, prioritized by stability (e.g., ID > role-based > text-based > CSS/XPath).
        If no URL was crawled, generate generic but reliable locators and insights based on common web patterns and the user's instruction.
        Ensure locators are:
        - Reliable and stable
        - Adaptable to dynamic content
        - Use self-healing strategies where possible
        - Include ID, name, class name, tag name, CSS selector, XPath, and role-based selectors
        - Use text-based selectors where applicable
        - Prioritize selectors based on reliability and stability
        """
        super().__init__("SiteInspector", system_message, model_name=DEFAULT_GROQ_MODEL)

    async def crawl_site(self, start_url: str, username: str, password: str, max_pages: int = 5) -> dict:
        """Simple BFS crawler to fetch up to max_pages internal pages and their HTML snippets after logging in."""
        from collections import deque
        visited = set()
        to_visit = deque([start_url])
        page_contents = {}
        base_origin = urlparse(start_url).scheme + "://" + urlparse(start_url).netloc

        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            context = await browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                viewport={"width": 1280, "height": 720}
            )
            page = await context.new_page()
            try:
                # Navigate to login page
                st.write(f"Navigating to {start_url}...")
                await page.goto(start_url, wait_until="domcontentloaded", timeout=60000)

                # Check if a "Sign In" button/link needs to be clicked
                sign_in_button = await page.query_selector("a[href*='login'], button:has-text('Sign In'), button:has-text('Log In')")
                if sign_in_button:
                    st.write("Clicking 'Sign In' button...")
                    await sign_in_button.click()
                    await page.wait_for_load_state("domcontentloaded", timeout=30000)

                # Selectors for email, password, and submit button. These can be expanded based on common patterns.
                email_selectors = [
                    "#userNameInput", "data-testid='email'", "input[type='email']", 
                    "input[name='email']", "input[id='email']", "//input[contains(@placeholder, 'Email')]"
                ]
                password_selectors = [
                    "#passwordInput", "data-testid='password'", "input[type='password']", 
                    "input[name='password']", "input[id='password']", "//input[contains(@placeholder, 'Password')]"
                ]
                submit_selectors = [
                    "#submitButton", ".submit", "[role='button']:has-text('Sign in')", 
                    "data-testid='submit'", "button[type='submit']", 
                    "button:has-text('Sign In')", "button:has-text('Log In')"
                ]

                email_locator = None
                for selector in email_selectors:
                    try:
                        await page.wait_for_selector(selector, state="visible", timeout=10000)
                        email_locator = selector
                        break
                    except:
                        continue

                if not email_locator:
                    html = await page.content()
                    st.error(f"Error: No email input found. Page HTML:\n{html[:1000]}...")
                    raise Exception("No email input found with provided selectors")

                st.write(f"Filling email with selector: {email_locator}")
                await page.fill(email_locator, username)
                await page.wait_for_timeout(1000) 

                password_locator = None
                for selector in password_selectors:
                    try:
                        await page.wait_for_selector(selector, state="visible", timeout=10000)
                        password_locator = selector
                        break
                    except:
                        continue

                if not password_locator:
                    raise Exception("No password input found with provided selectors")

                st.write(f"Filling password with selector: {password_locator}")
                await page.fill(password_locator, password)
                await page.wait_for_timeout(1000)

                submit_locator = None
                for selector in submit_selectors:
                    try:
                        await page.wait_for_selector(selector, state="visible", timeout=10000)
                        submit_locator = selector
                        break
                    except:
                        continue

                if not submit_locator:
                    raise Exception("No submit button found with provided selectors")

                st.write(f"Clicking submit with selector: {submit_locator}")
                await page.click(submit_locator)

                # Wait for post-login page
                try:
                    await page.wait_for_selector(".search-panel, #searchPanel, [role='search']", state="visible", timeout=30000)
                    st.write(f"Logged in successfully at {start_url}")
                except:
                    error_selector = "text='Invalid credentials', text='Login failed', [role='alert']"
                    error_element = await page.query_selector(error_selector)
                    if error_element:
                        error_text = await error_element.inner_text()
                        st.error(f"Login failed with error: {error_text}")
                        raise Exception(f"Login failed: {error_text}")
                    await page.wait_for_timeout(5000)
                    current_url = page.url
                    if current_url == start_url:
                        html = await page.content()
                        st.error(f"Error: No redirect after login. Current URL: {current_url}\nPage HTML:\n{html[:1000]}...")
                        raise Exception("No redirect after login attempt")
                    st.write(f"Redirected to {current_url} after login")

                # Start crawling after login
                while to_visit and len(page_contents) < max_pages:
                    current = to_visit.popleft()
                    if current in visited:
                        continue
                    visited.add(current)
                    try:
                        st.write(f"Crawling page: {current}")
                        await page.goto(current, wait_until="domcontentloaded", timeout=60000)
                        html = await page.content()
                        page_contents[current] = html[:4000] # Snippet
                        
                        new_links = await page.evaluate('''
                            (base_origin) => {
                                return Array.from(document.querySelectorAll('a[href]'))
                                    .map(a => {
                                        let href = a.getAttribute('href');
                                        if (href) {
                                            try {
                                                let fullUrl = new URL(href, window.location.href).href;
                                                if (fullUrl.startsWith(base_origin)) {
                                                    return fullUrl;
                                                }
                                            } catch (e) {}
                                        }
                                        return null;
                                    })
                                    .filter(Boolean);
                            }
                        ''', base_origin)

                        for link in new_links:
                            parsed = urlparse(link)
                            if (link not in visited and
                                link not in to_visit and
                                not any(link.lower().endswith(ext) for ext in ('.pdf', '.jpg', '.png', '.gif', '.css', '.js', '.zip')) and
                                parsed.path != '/' and parsed.path != ''):
                                to_visit.append(link)
                    except Exception as e:
                        st.write(f"Error crawling {current}: {e}")
                        continue
            except Exception as e:
                st.error(f"Error during login or crawling: {e}")
                if not page_contents:
                    try:
                        html = await page.content()
                        st.write(f"Page HTML on failure:\n{html[:1000]}...")
                    except Exception as page_e:
                        st.error(f"Could not even get page content on failure: {page_e}")
            finally:
                await context.close()
                await browser.close()
        return page_contents

    async def inspect_site(self, url: str, key_elements: str, instruction: str, username: str, password: str) -> str:
        if url:
            if not username or not password:
                st.warning("Username or password not provided in prompt. Crawling without login.")
                # Implement a non-login crawl or return generic response
                return await self.generate_response(
                    f"No login credentials provided. Generate reliable Playwright locators and insights for {key_elements} based on common web patterns and the instruction: {instruction}"
                )
                
            page_contents = await self.crawl_site(url, username, password, max_pages=5)
            if not page_contents:
                st.warning("No pages crawled successfully. Generating generic insights.")
                return await self.generate_response(
                    f"No URL content crawled. Generate reliable Playwright locators and insights for {key_elements} based on common web patterns and the instruction: {instruction}"
                )
            content_str = "\n\n---\n\n".join([f"Page: {k}\nHTML Snippet:\n{v}" for k, v in page_contents.items()])
            crawl_summary = await self.generate_response(
                f"Start URL: {url}\nKey Elements to Focus: {key_elements}\nUser Instruction: {instruction}\nCrawled Pages Snippets:\n{content_str}"
            )
            recommendations = await self.generate_response(
                f"Analyze the crawl summary for site insights and locators: {crawl_summary}\nUser Key Elements: {key_elements}\nUser Instruction: {instruction}"
            )
            return recommendations
        else:
            st.warning("No URL provided. Generating generic insights.")
            return await self.generate_response(
                f"No URL provided. Generate reliable Playwright locators, self-healing strategies, and generic site insights (e.g., common flows for {key_elements}) based on common web patterns and the instruction: {instruction}"
            )

class PlannerAgentOSS(GroqOSSAgent):
    def __init__(self):
        system_message = """
        You are an expert QA test planner with deep NLP understanding.
        Your goal is to generate comprehensive test cases covering all possible variations, including but not limited to:

        Firstly, your priority is to generate test cases for the core functionalities described in the instruction, including insights from crawled site data.
        The core functionalities include covering complete flows for each feature mentioned in the instruction and discovered during site crawling.
        The features are the basic flow, alternative flow, pre-conditions, post-conditions, validations/rules mentioned in the instruction, and additional scenarios from crawled data.
        Then, expand to cover edge cases, error handling, and less common scenarios.
        After completing the core functionalities, generate test cases for the following types:
        - Functional (positive scenarios where the system works as expected)
        - Negative (invalid inputs, error handling, failures)
        - Boundary (edge cases like min/max values, limits)
        - Performance (load times, responsiveness under stress; simulate with Playwright where possible, e.g., multiple interactions, timeouts)
        - Security (vulnerabilities like injection, authentication bypass; automate checks for common issues like XSS, CSRF if detectable via UI)
        - Integration (interactions between components, APIs if accessible via UI)
        - Usability (UI/UX checks like accessibility, responsiveness, user flows; use Playwright for visibility, focus, etc.)
        - Regression (re-testing core functionalities to ensure no breaks)
        - Smoke (basic functionality checks to verify build stability)
        - Sanity (quick checks on specific changes or fixes)
        - Database (if applicable, verify data persistence, queries via UI interactions)
        - End-to-End (full user journeys from start to finish)
        - Exploratory (suggest automated heuristics or random inputs for discovery; adapt to automation where feasible)

        Analyze the provided instruction, refined details, and site insights/locator recommendation to generate test cases for as many of these types as applicable. If a type doesn't apply, skip it but aim to cover all possible variations where relevant.
        Prioritize generating multiple test cases per type to cover variations (e.g., different inputs, scenarios).
        For each test case, include:
        Strictly use this criteria for the generated test cases:
        - Test Case ID: A unique identifier in the format TC-<number> (e.g., TC-1, TC-2) 
        - High Level Feature (e.g Login, Search etc)
        - Feature Name (Specific feature under the high-level feature, e.g., Login with valid credentials)
        - Test Scenario (High-level overview of the functionality being tested)
        - Test Case  (This includes the details of the functionality being tested, including all possible variations. High-level overview)
        - Test Case Description (Detailed description of what the test case aims to validate)
        - Step-by-step actions with clear selectors, actions, and validations (Use Playwright sync API, self-healing locators, waits, per-step logging/assertions)
        - Possible Values (if applicable, e.g., input data variations)
        - Sources (Any system or Database used to confirm expected results if applicable. Where this is not applicable, write 'N/A') 
        - Expected Result (Clear pass/fail criteria)
        - Data Correctness Checked (Yes/No if applicable. If not applicable, write 'N/A')
        - Release/Platform Version (Web/Mobile/IOS/Android etc. If not applicable, write 'N/A')
        - Automation Possibility (Yes/No)
        - Testing Type (Indicate type, e.g., Functional, Negative, Boundary, Performance, Security, Integration, Usability, Regression, Smoke, Sanity, Database, End-to-End, Exploratory)
        - Priority (High, Medium, Low)
        - Testing Phase : QA 
        ALWAYS WITH EACH TEST CASE:
        OUTPUT the test cases in the following format:
        STRICTLY ADHERE TO THIS FORMAT:
        - Test Case ID: TC-<number>
        - High Level Feature
        - Feature Name
        - Test Scenario
        - Test Case
        - Test Case Description
        - Step-by-step actions
        - Possible Values (if applicable, Type 'None' if there is none for a specific case)
        - Sources (if applicable, Type 'N/A' if there is none for a specific case)
        - Expected Result
        - Data Correctness Checked (if applicable, Type 'N/A' if there is none for a specific case)
        - Release/Platform Version (Web/Mobile/IOS/Android etc. If not applicable, write 'N/A')
        - Automation Possibility
        - Testing Type
        - Priority
        - Testing Phase : QA
        Guidelines:
        
        1.  Use the exact sub-headings: `Test Case ID`, `High Level Feature`, `Feature Name`, `Test Scenario`,`Test Case`,`Test Case Description`, `Step-by-step actions`,`Possible Values`, `Sources`, `Expected Result`, `Data Correctness Checked`, `Release/Platform Version`, `Automation Possibility`, `Testing Type`, `Priority` and `Testing Phase`.
        2.  Number the Test Case ID sequentially starting from 1 (e.g., TC-1, TC-2, etc.).
        3. The fields `Test Case ID`, `High Level Feature`, `Feature Name`, `Test Scenario`,`Test Case`,`Test Case Description`, `Possible Values`, `Sources`, `Expected Result`, `Data Correctness Checked`, `Release/Platform Version`, `Automation Possibility`, `Testing Type`, and `Priority` MUST be single lines using the bullet (`*`) prefix.
        4. The steps under `Step-by-step actions` should not be a numbered list, instead a paragraph with all steps in a sequence without any numbering or bullet points.
        5. The step by step actions or any fields of testcase should NOT include locator reccomendation itself. 
        Please always output the testcases as I described above.

        Structure your response with sections for each test type (e.g., ## Functional Test Cases, ## Negative Test Cases, etc.).
        Under each section, provide a numbered list of test cases.
        Use precise language and avoid ambiguity.
        Focus on:
        - Setup steps (navigate, prepare data)
        - Action steps (click, fill, submit)
        - Verification steps (assertions, checks)
        - Error handling considerations
        - Use clear, actionable language
        - Output only the test cases, no explanations or markdown beyond the required section headers and numbered lists
        - Follow Playwright sync API conventions
        - Use self-healing locator strategies (e.g., ID, name, class name, tag name, CSS selector, XPath, role-based selectors, text-based selectors)
        - Prioritize selectors based on reliability and stability
        - Include self-healing locator strategies (e.g., role-based, text-based over IDs if dynamic)
        - Ensure each test case is executable with clear pass/fail criteria
        - Include per-step pass/fail logging and assertions (e.g., console.log('Step 1: Passed') or expect().toBeVisible())
        - Use the provided instruction, refined details, and locator recommendations/site insights as context for generating test cases
        - For performance/security/usability, adapt to Playwright capabilities (e.g., measure page load time, check for alerts, verify ARIA attributes)
        - For exploratory, generate test cases with randomized or varied inputs to simulate exploration
        - Generate only the test cases in this step, dont output any refined instruction, explanations, or locator recommendations.
        - First generate test cases for core functionalities, including those derived from site crawling insights, covering basic flow, alternate flow, pre-conditions, post-conditions, validations/rules mentioned in the instruction.
        - Then expand to cover all other types of test cases as mentioned above.
        - Dont include the locator reccomendation itself in the description of testcases fields, but instead use the insights from crawled data. Dont mention any locators in testcases.
        """
        super().__init__("PlannerOSS", system_message, model_name=DEFAULT_GROQ_MODEL)

class UserProxyAgent:
    def __init__(self, name: str):
        self.name = name

    async def initiate_chat(self, agent, message: str) -> str:
        return await agent.generate_response(message)

# --- 2. MODIFIED PARSING FUNCTION (Added st.success/warning) ---


def parse_and_export_testcases(test_cases_str: str):
    """
    Parses test cases from a string with variable formatting and exports them to an Excel file.
    """
    # Add re.MULTILINE and anchors (^) to only split on titles at the start of a line
    test_matches = re.findall(r'(\d*)\.?\s*\*\*(.+?)\*\*\s*((?:.|\n)*?)(?=\n\s*\d*\.?\s*\*\*|\Z)', test_cases_str)
    all_data = []
    st.session_state.test_cases_list = []  # Reset test cases list

    # Fallback regex if the primary one fails (e.g., no markdown bolding)
    if not test_matches:
        test_matches = re.findall(
            r'(\d+)\.\s*(.+?)\n((?:.|\n)*?)(?=\n\s*\d+\.|\Z)',
            test_cases_str
        )

    for num, High_Level_Feature, content in test_matches:
        data = {
            'Test Case ID': '',
            'High Level Feature': High_Level_Feature.strip(),
            'Feature Name': '',
            'Test Scenario': '',
            'Test Case': '',
            'Test Case Description': '',
            'Step-by-step actions': '',
            'Possible Values': '',
            'Sources': '',
            'Expected Result': '',
            'Data Correctness Checked': '',
            'Release/Platform Version': '',
            'Automation Possibility': '',
            'Testing_Type': '',
            'Priority': '',
            'Testing Phase': ''
        }

        # Try to find the keys. Relaxed the regex to not require '-'
        # field_patterns = {
        #     'Test Case ID': r'(?:[-*]\s*)?Test\s*Case\s*ID\s*:\s*(.+?)(?=\n\s*[-*]|\Z)',
        #     'Title': r'[o\*-]\s*(?:\*\*)?Title(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
        #     'Test Scenario': r'[o\*-]\s*(?:\*\*)?Test\s*Scenario(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',   
        #     'Testing_Type': r'[o\*-]\s*(?:\*\*)?Testing\s*Type(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
        #     'Test Case': r'[o\*-]\s*(?:\*\*)?Test\s*Case(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
        #     'Step-by-step actions': r'[o\*-]\s*(?:\*\*)?Step-by-step\s*actions(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
        #     'Possible Values': r'[o\*-]\s*(?:\*\*)?Possible\s*Values(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
        #     'Expected Result': r'[o\*-]\s*(?:\*\*)?Expected\s*Result(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
        # }
        field_patterns = {
    'Test Case ID': r'(?:[-*]\s*)?Test\s*Case\s*ID\s*:\s*(.+?)(?=\n\s*[-*]|\Z)',
    'High Level Feature': r'[o\*-]?\s*(?:\*\*)?High\s*Level\s*Feature(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Feature Name': r'[o\*-]?\s*(?:\*\*)?Feature\s*Name(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Test Scenario': r'[o\*-]\s*(?:\*\*)?Test\s*Scenario(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Test Case': r'[o\*-]\s*(?:\*\*)?Test\s*Case(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Test Case Description': r'[o\*-]?\s*(?:\*\*)?Test\s*Case\s*Description(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Step-by-step actions': r'[o\*-]\s*(?:\*\*)?Step-by-step\s*actions(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Possible Values': r'[o\*-]\s*(?:\*\*)?Possible\s*Values(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Sources': r'[o\*-]?\s*(?:\*\*)?Sources(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Expected Result': r'[o\*-]\s*(?:\*\*)?Expected\s*Result(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Data Correctness Checked': r'[o\*-]?\s*(?:\*\*)?Data\s*Correctness\s*Checked(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Release/Platform Version': r'[o\*-]?\s*(?:\*\*)?Release\/Platform\s*Version(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Automation Possibility': r'[o\*-]?\s*(?:\*\*)?Automation\s*Possibility(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Testing_Type': r'[o\*-]\s*(?:\*\*)?Testing\s*Type(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Priority': r'[o\*-]?\s*(?:\*\*)?Priority(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',
    'Testing Phase': r'[o\*-]?\s*(?:\*\*)?Testing\s*Phase(?:\*\*)?\s*:\s*(.+?)(?=\n\s*[o\*-]|\Z)',

}


        for key, pattern in field_patterns.items():
            match = re.search(pattern, content, re.DOTALL | re.IGNORECASE)
            if match:
                value = match.group(1).strip()
                value = re.sub(r'\*\*|_+|`+', '', value)  # Clean markdown
                if key == "Step-by-step actions":
                    value = re.sub(r'\n\s*(\d+\.)\s*', r'\n\1 ', value)
                    value = re.sub(r'^\s*(\d+\.)\s*', r'\1 ', value, flags=re.MULTILINE)
                else:
                    value = re.sub(r'\s*\n\s*', ' ', value)
                data[key] = value.strip()

        all_data.append(data)
        st.session_state.test_cases_list.append(data)  # Store in session state

    output_path = "cleaned_generated_test_cases.xlsx"
    if all_data:
        df = pd.DataFrame(all_data)
        
        try:
            df.to_excel(output_path, index=False)

            # Format Excel cells
            wb = load_workbook(output_path)
            ws = wb.active
            for col in ws.columns:
                max_length = 0
                column = get_column_letter(col[0].column)
                for cell in col:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        max_length = max(max_length, *[len(line) for line in lines])
                
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = min(adjusted_width, 70)

            wb.save(output_path)
            st.success(f"‚úÖ Test cases exported successfully to {output_path}")
        except Exception as e:
            st.error(f"Error saving or formatting Excel file: {e}")
    else:
        st.warning("‚ö†Ô∏è No test cases were parsed ‚Äî please check the LLM output format.")


# --- 3. STREAMLIT UI & ASYNC LOGIC ---

# Page config
st.set_page_config(page_title="ü§ñ QA Test Case Generator", layout="wide")

# Initialize Agents (cached to avoid re-creation on every rerun)
@st.cache_resource
def get_agents():
    return {
        "user": UserProxyAgent("User"),
        "inspector": SiteInspectorAgent(),
        "planner": PlannerAgentOSS()
    }

agents = get_agents()
user = agents["user"]
inspector = agents["inspector"]
planner = agents["planner"]

# --- Async Helper Functions ---

async def run_initial_generation(user_prompt, status_placeholder):
    """Orchestrates the full initial generation process."""
    
    # 1. Extract details from prompt
    status_placeholder.update(label="Extracting details from prompt...")
    url_match = re.search(r'(https?://[^\s]+)', user_prompt)
    site_url = url_match.group(1) if url_match else None
    username_match = re.search(r"username\s*=\s*'([^']+)'", user_prompt)
    password_match = re.search(r"password\s*=\s*'([^']+)'", user_prompt)
    username = username_match.group(1) if username_match else None
    password = password_match.group(1) if password_match else None
    element_keywords = [kw for kw in ["search", "input", "button", "link", "verify", "assert", "click", "fill", "submit", "navigate", "page", "text", "selector"] if kw in user_prompt.lower()]
    key_elements = ", ".join(element_keywords) if element_keywords else "main interactive elements"

    # 2. Refine Instruction
    status_placeholder.update(label="Step 1/3: Refining instruction...")
    refined = await refine_instruction(user_prompt)
    st.session_state.refined_instruction = refined

    # 3. Inspect Site
    crawl_status_container = st.container()
    with crawl_status_container:
        status_placeholder.update(label=f"Step 2/3: Inspecting {site_url or 'site'}... (This may take a moment)")
        locators = await inspector.inspect_site(site_url, key_elements, refined, username, password)
    st.session_state.locator_recommendations = locators
    crawl_status_container.empty() # Clear crawl messages

    # 4. Plan Test Cases
    planner_input = f"Refined Instruction:\n{refined}\n\nSite Insights and Recommended Locators:\n{locators}"
    st.session_state.planner_input = planner_input # Save for feedback
    
    status_placeholder.update(label="Step 3/3: Planning initial test cases...")
    initial_cases = await user.initiate_chat(planner, planner_input)
    st.session_state.all_test_cases_str = initial_cases

    # 5. Parse and Save
    status_placeholder.update(label="Parsing and saving test cases to Excel...")
    parse_and_export_testcases(st.session_state.all_test_cases_str)


async def run_feedback_generation(feedback_prompt, status_placeholder):
    """Orchestrates the feedback-based generation process."""
    
    # 1. Create new prompt for planner
    status_placeholder.update(label="Formulating feedback for planner...")
    planner_input = (
        f"You have already generated a set of test cases. A human user now wants you to ADD MORE test cases based on the following feedback. "
        f"**IMPORTANT: Generate ONLY the NEW test cases requested. Do NOT repeat or modify the previously generated ones.**\n\n"
        
        f"**STRICTLY ADHERE TO THE FOLLOWING OUTPUT FORMAT FOR NEW CASES:**\n"
        f"- Test Case ID: TC-<number> (continue numbering from previous cases)\n"
        f"- High Level Feature\n"
        f"- Feature Name\n"
        f"- Test Scenario\n"
        f"- Test Case\n"
        f"- Test Case Description\n"
        f"- Step-by-step actions\n"
        f"- Possible Values (if applicable, Type 'None' if there is none for a specific case)\n"
        f"- Sources (if applicable, Type 'N/A' if there is none for a specific case)\n"
        f"- Expected Result\n"
        f"- Data Correctness Checked (if applicable, Type 'N/A' if there is none for a specific case)\n"
        f"- Release/Platform Version (if applicable, Type 'N/A' if there is none for a specific case)\n"
        f"- Automation Possibility\n"
        f"- Testing Type\n"
        f"- Priority\n\n"
        f"- Testing Phase\n: QA\n"
        f"**GUIDELINES:**\n"
        f"- Use the exact sub-headings: `Test Case ID`, `High Level Feature`, `Feature Name`, `Test Scenario`,`Test Case`,`Test Case Description`, `Step-by-step actions`,`Possible Values`, `Sources`, `Expected Result`, `Data Correctness Checked`, `Release/Platform Version`, `Automation Possibility`, `Testing Type`, `Priority` and `Testing Phase`.\n"
        f"- The fields `Test Case ID`, `High Level Feature`, `Feature Name`, `Test Scenario`,`Test Case`,`Test Case Description`, `Possible Values`, `Sources`, `Expected Result`, `Data Correctness Checked`, `Release/Platform Version`, `Automation Possibility`, `Testing Type`, and `Priority` MUST be single lines using the bullet (`*`) prefix.\n"
        f"- Number the Test Case ID sequentially starting from the last generated Test Case ID.\n"
        f"- The steps under `Step-by-step actions` should not be a numbered list, instead a paragraph with all steps in a sequence without any numbering or bullet points.\n"
        f"- **Generate ONLY the new test cases requested in the feedback.**\n\n"
        f"Feedback: '{feedback_prompt}'"
    )
        
    
    # 2. Generate new test cases
    status_placeholder.update(label="Generating additional test cases...")
    new_test_cases = await user.initiate_chat(planner, planner_input)
    
    # 3. Append and save
    status_placeholder.update(label="Appending new test cases...")
    # The new test cases are appended to the existing string
    if st.session_state.all_test_cases_str:
        st.session_state.all_test_cases_str += "\n\n" + new_test_cases
    else:
        st.session_state.all_test_cases_str = new_test_cases
    
    # 4. Re-parse and save the CUMULATIVE list
    status_placeholder.update(label="Parsing and saving all test cases to Excel...")
    parse_and_export_testcases(st.session_state.all_test_cases_str)
# --------------------------------------------------------------
# 1. EDIT FUNCTION (replace your current run_edit_generation)
# --------------------------------------------------------------
import re
import streamlit as st


def parse_test_case_block(block: str):
    """
    Extract:
    - fields as dict
    - original prefix (*, ‚Ä¢, -, etc.)
    - optional numeric heading (e.g., '2.')
    """
    fields = {}
    prefix = "*"
    numeric_header = None

    lines = block.split("\n")

    for line in lines:
        stripped = line.strip()

        # Detect numeric header like "2."
        num_match = re.match(r'^(\d+)\.$', stripped)
        if num_match:
            numeric_header = num_match.group(1)
            continue

        # Detect field lines
        field_match = re.match(r'^([*‚Ä¢-])\s*(.+?):\s*(.*)', stripped)
        if field_match:
            prefix = field_match.group(1)
            key = field_match.group(2).strip()
            value = field_match.group(3).strip()
            fields[key] = value

    return fields, prefix, numeric_header



# -----------------------------
# Helper: Rebuild block from dict
# -----------------------------
def rebuild_test_case_block(fields: dict, prefix: str, numeric_header: str):
    rebuilt = ""

    # Restore numbering like "2."
    if numeric_header:
        rebuilt += f"{numeric_header}.\n\n"

    # Restore bullet style
    for key, value in fields.items():
        rebuilt += f"{prefix} {key}: {value}\n"

    return rebuilt.strip()



# -----------------------------
# Helper: Detect target field
# -----------------------------
async def detect_target_field(edit_instruction: str, fields: dict):
    field_list = list(fields.keys())

    detection_prompt = f"""
You are a QA assistant.

Given this edit instruction:
"{edit_instruction}"

And this list of available fields:
{field_list}

Return ONLY the exact field name that should be edited.
Do not return anything else.
"""

    response = await user.initiate_chat(planner, detection_prompt)
    return response.strip()


# -----------------------------
# MAIN FUNCTION (UPDATED)
# -----------------------------
async def run_edit_generation(test_case_id: str, edit_instruction: str, status_placeholder):

    # Normalize text
    st.session_state.all_test_cases_str = (
        st.session_state.all_test_cases_str
        .replace("\r\n", "\n")
        .replace("\r", "\n")
        .strip() + "\n"
    )

    status_placeholder.update(label="üîç Searching for test case...")

    pattern = rf'(\d+\.\s*\n\s*\n)?(\s*[*‚Ä¢-]\s*Test Case ID:\s*{re.escape(test_case_id.strip())}.*?)(?=\n\d+\.|\Z)'

    match = re.search(pattern, st.session_state.all_test_cases_str, re.DOTALL)

    if not match:
        st.error(f"‚ùå Test case {test_case_id} not found.")
        return None, None

    original_block = match.group(0).strip()

    # Parse block safely
    fields, prefix, numeric_header = parse_test_case_block(original_block)

    # Detect which field to edit
    status_placeholder.update(label="üß† Detecting field to edit...")
    target_field = await detect_target_field(edit_instruction, fields)

    if target_field not in fields:
        st.error(f"‚ùå Invalid field detected: {target_field}")
        return None, None

    original_value = fields[target_field]

    # Ask LLM to regenerate ONLY that field value
    status_placeholder.update(label=f"‚úèÔ∏è Updating: {target_field}")

    field_prompt = f"""
You are a strict QA editor.

Instruction:
"{edit_instruction}"

Field Name:
{target_field}

Original Value:
{original_value}

Return ONLY the updated value.
Do NOT include field name.
Do NOT include bullets.
Do NOT include extra text.
"""

    updated_value = await user.initiate_chat(planner, field_prompt)
    updated_value = updated_value.strip()

    # Update only target field
    fields[target_field] = updated_value

    # Rebuild block preserving numbering + bullets
    updated_block = rebuild_test_case_block(fields, prefix, numeric_header)

    # Replace safely
    status_placeholder.update(label="üß© Replacing test case...")

    new_full_str, n_subs = re.subn(
        pattern,
        updated_block,
        st.session_state.all_test_cases_str,
        count=1,
        flags=re.DOTALL
    )

    if n_subs == 0:
        st.error("‚ùå Replacement failed.")
        return None, None

    # st.session_state.all_test_cases_str = (
    #     new_full_str.replace("\r\n", "\n").replace("\r", "\n").strip() + "\n"
    # )
    st.session_state.all_test_cases_str = (
    new_full_str.replace("\r\n", "\n").replace("\r", "\n").strip() + "\n"
)

# üî• IMPORTANT: Regenerate Excel file with updated data
    try:
        parse_and_export_testcases(st.session_state.all_test_cases_str)
    except Exception as e:
        st.warning(f"‚ö†Ô∏è Excel regeneration failed after edit: {e}")


    status_placeholder.update(label="‚úÖ Update successful", state="complete")
    st.success(f"‚úÖ Test case {test_case_id} updated successfully!")

    return original_block, updated_block




# --- Streamlit UI ---

st.title("ü§ñ Multi-Agent QA Test Case Generator")
st.markdown("This tool uses a team of AI agents (powered by Groq) to crawl a website, analyze your instructions, and generate comprehensive QA test cases.")

# Initialize session state
if 'all_test_cases_str' not in st.session_state:
    st.session_state.all_test_cases_str = ""
if 'refined_instruction' not in st.session_state:
    st.session_state.refined_instruction = ""
if 'locator_recommendations' not in st.session_state:
    st.session_state.locator_recommendations = ""

# --- Sidebar for Inputs ---
with st.sidebar:
    st.image("https://avatars.githubusercontent.com/u/153243936?s=200&v=4", width=100)
    st.header("1. User Instruction")
    st.markdown("Provide your test instruction, including URL and credentials.")
    
    example_prompt = "Test the login flow at https://example.com with username='user@test.com' and password='password123'. Verify successful login by checking for the 'Dashboard' text."
    user_prompt = st.text_area("Your Instruction:", value=example_prompt, height=150, key="user_prompt_input")
    
    generate_button = st.button("üöÄ Generate Initial Test Cases", use_container_width=True, type="primary")

    st.divider()
    
    st.header("2. Feedback Loop")
    st.markdown("Review the generated cases. If you need more, provide feedback below.")
    feedback = st.text_area("Feedback (e.g., 'add tests for the forgot password link'):", height=100, key="feedback_input")
    
    feedback_button = st.button("üîÑ Generate More Cases", use_container_width=True, type="primary")

    st.divider()
    # st.info(f"Using model: **{DEFAULT_GROQ_MODEL}**")

    st.divider()
    st.header("3. Edit Test Case via Prompt")
    st.markdown("Tell the AI how to update a specific test case. Only that case will be changed.")

    edit_id = st.text_input("Test Case ID to Edit (e.g., TC-51):", key="edit_id_input")
    edit_prompt = st.text_area(
        "Edit Instruction (natural language):",
        placeholder="e.g., Update TC-51: Add step to verify no duplicate suggestions and check empty state",
        height=120,
        key="edit_prompt_input"
    )
    edit_button = st.button("Update Test Case", use_container_width=True, type="primary")

    # Optional: Show current test case
    # --------------------------------------------------------------
# 2. UI BLOCK ‚Äì CALLING THE EDIT FUNCTION
# --------------------------------------------------------------
# --- Main Area for Outputs ---
output_container = st.container()

# Button Clicks & Async Logic
if generate_button and user_prompt:
    # Reset state for a new run
    st.session_state.all_test_cases_str = ""
    st.session_state.refined_instruction = ""
    st.session_state.locator_recommendations = ""
    
    with st.status("üöÄ Starting test generation process...", expanded=True) as status:
        try:
            asyncio.run(run_initial_generation(user_prompt, status))
            status.update(label="‚úÖ Generation complete!", state="complete")
        except Exception as e:
            st.error(f"An error occurred during generation: {e}")
            status.update(label="Generation failed.", state="error")
    st.rerun() # Rerun to update the main display

if feedback_button and feedback:
    if not st.session_state.all_test_cases_str:
        st.warning("Please generate initial test cases first before providing feedback.")
    else:
        with st.status("üîÑ Incorporating feedback...", expanded=True) as status:
            try:
                asyncio.run(run_feedback_generation(feedback, status))
                status.update(label="‚úÖ Additional cases generated!", state="complete")
            except Exception as e:
                st.error(f"An error occurred during feedback generation: {e}")
                status.update(label="Feedback generation failed.", state="error")
        # Clear feedback box by re-running
        st.rerun()

if edit_button and edit_id and edit_prompt:
    if not st.session_state.all_test_cases_str:
        st.warning("‚ö†Ô∏è Please generate test cases first before editing.")
    else:
        with st.status(f"‚úèÔ∏è Editing **{edit_id}**...", expanded=True) as status:
            try:
                raw_llm, cleaned = asyncio.run(
                    run_edit_generation(edit_id.strip(), edit_prompt, status)
                )

                # ---- ALWAYS SHOW RAW LLM OUTPUT ----
                with st.expander("üîç DEBUG: Raw LLM Output", expanded=True):
                    if raw_llm:
                        st.code(raw_llm, language="markdown")
                        st.caption(f"Output length: {len(raw_llm)} characters")
                    else:
                        st.warning("No output received from LLM.")

                # ---- SHOW CLEANED BLOCK IF EXTRACTED ----
                if cleaned:
                    with st.expander("‚úÖ Cleaned Output Used for Replacement", expanded=False):
                        st.code(cleaned, language="markdown")
                        st.caption(f"Cleaned length: {len(cleaned)} characters")
                    
                    status.update(label=f"‚úÖ {edit_id} updated successfully!", state="complete")
                    st.success(f"Test case **{edit_id}** has been updated!")
                    
                    # Show before/after comparison
                    if hasattr(st.session_state, 'all_test_cases_str_before_edit'):
                        with st.expander("üìä Before/After Comparison", expanded=False):
                            col1, col2 = st.columns(2)
                            with col1:
                                st.markdown("**Before:**")
                                # Extract the old version
                                pattern = rf'([*‚Ä¢]\s*Test Case ID:\s*{re.escape(edit_id.strip())}\s*\n.*?)(?=\n[*‚Ä¢]\s*Test Case ID:|\Z)'
                                old_match = re.search(pattern, st.session_state.all_test_cases_str_before_edit, re.DOTALL)
                                if old_match:
                                    st.code(old_match.group(1).strip(), language="markdown")
                            with col2:
                                st.markdown("**After:**")
                                st.code(cleaned, language="markdown")
                    
                    # st.rerun()
                else:
                    st.error("‚ùå Could not extract a valid test case block from the AI response.")
                    st.info("Check the raw LLM output above to see what the AI returned.")
                    status.update(label="‚ö†Ô∏è Edit failed - check raw output", state="error")

            except Exception as e:
                st.error(f"‚ùå Unexpected error during edit: {e}")
                import traceback
                with st.expander("üêõ Full Error Traceback", expanded=True):
                    st.code(traceback.format_exc())
                status.update(label="‚ùå Edit failed", state="error")

# --- Display Results ---
with output_container:
    if not st.session_state.refined_instruction and not st.session_state.all_test_cases_str:
        st.markdown("### Welcome! üëã")
        st.markdown("Enter your test instruction in the sidebar and click **'Generate Initial Test Cases'** to begin.")
        st.markdown("Your results will appear here.")
        
    if st.session_state.refined_instruction:
        with st.expander("Step 1: Refined Instruction", expanded=False):
            st.markdown(st.session_state.refined_instruction)
    
    if st.session_state.locator_recommendations:
        with st.expander("Step 2: Site Insights & Locator Recommendations", expanded=False):
            st.markdown(st.session_state.locator_recommendations)

    if st.session_state.all_test_cases_str:
        st.header("Step 3: Cumulative Generated Test Cases")
        
        # Add the download button
        try:
            with open("cleaned_generated_test_cases.xlsx", "rb") as fp:
                st.download_button(
                    label="üì• Download All Test Cases (Excel)",
                    data=fp,
                    file_name="generated_test_cases.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        except FileNotFoundError:
            st.error("Excel file ('cleaned_generated_test_cases.xlsx') not found. Please try generating again.")
        
        # Display the raw test cases

        st.markdown(st.session_state.all_test_cases_str)







